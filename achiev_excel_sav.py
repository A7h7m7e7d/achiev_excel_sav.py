import openpyxl as ex
from datetime import date
file= ex.load_workbook("date.xlsx")
sheet=file.active
sheet=file["Sheet1"]

def num_row(col,sheet):
    column_number = col

    column_data = []
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=column_number, max_col=column_number):
        for cell in row:
            column_data.append(cell.value)
    fulcolumn_data=[]
    for item in column_data:
        if item != None:
            fulcolumn_data.append(item)
    return len(fulcolumn_data)




















def onlyfil(x):
    c=0
    for i in range(x.max_column):

        if str(x[1][i].value)!= str(None):

            c+=1
    return c
def onlyfilr(x):
    c=0
    for i in range(x[onlyfil(x)].max_row):

        if str(x[1][i].value)!= str(None):

            c+=1
    return c
def remove(x):
    d=[]
    for i in x:
       if i != ";":
           d.append(i)
    return ''.join(d)
def choice():
    ch = input("what is your choce 1 or 2 or 3 ?")

    while ch.strip() !="1" and ch!="2" and ch=="3" :

        ch = input("what is your choce 1 or 2  or 3 ?")
    return ch



def whatdo():
    print("If you want to read what did you do this day enter (1)")
    print("if you want to add a new achievement enter (2)")
    print("exit enter 3")
    d=choice()
    if d=="1":
        read()
    elif d=="3":
        return 0

    else:add()

def read():
    d = date.today()
    d=str(d)
    dat=input("what is the day: format be like "+ d +" ")
    for col in sheet.iter_cols():


        if str(col[0].value) == dat:

            for cell in col:
                print(cell.value)
m=onlyfil(sheet)
def add():
    d= date.today()
    m=onlyfil(sheet)
 




    
    if remove(str(sheet[1][m-1].value)) != str(d):

        sheet[1][m].value=str(d)
        achiev = input(" what do you want to add? ")

        r=num_row(m+1,sheet)
        sheet[r+1][m].value=achiev

        file.save("date.xlsx")

    else:
        achiev= input(" what do you want to add? ")

        r1=num_row(m,sheet)
        r1 = r1 + 1
        print("rows ", r1, "colomes ", m)
        sheet[r1][m-1].value=achiev

        file.save("date.xlsx")
    while True:
        print("if you want to add something enter 1")
        d=choice()
        if d == "1" :
            add()
        else:break




#sheet.cell(row=1, column=1, value = d)
while True:


    whatdo()

    file.save("date.xlsx")





